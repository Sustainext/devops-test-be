import sys
import traceback
from io import BytesIO

from django.forms.models import model_to_dict
from django.http import HttpResponse
from django.core.files.storage import default_storage

from openpyxl import load_workbook

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from sustainapp.Serializers.CheckOrgCorpSerializer import CheckOrgCoprSerializer
from canadabills211.models.CanadaBillS211 import IdentifyingInformation, AnnualReport
from canadabills211.Serializers.CanadaBillS211Serializer import (
    IISerializer,
    ARSerializer,
)


class GenerateExcel(APIView):

    def validate_identifying_info(self, data):
        # Mandatory fields for Identifying Information
        mandatory_fields = [
            "report_purpose_1",
            "reporting_legal_name_2",
            "financial_reporting_year_from_3",
            "financial_reporting_year_to_3",
            "is_revised_version_4",
            "is_joint_report_6",
            "subject_to_supply_chain_legislation_7",
            "categorizations_8",
            "sectors_or_industries_9",
            "country_10",
        ]

        additional_fields = {
            "is_revised_version_4": [
                "Yes",
                ["original_report_date_4_1", "changes_description_4_2"],
            ],
            "is_joint_report_6": ["Yes", ["legal_name_and_business_numbers_6_1_2"]],
            "subject_to_supply_chain_legislation_7": ["Yes", ["applicable_laws_7_1"]],
            "applicable_laws_7_1": [
                "Other",
                ["other_laws_description_7_1"],
            ],
            "sectors_or_industries_9": [
                "Other",
                ["sectors_or_industries_description_9"],
            ],
            "country_10": ["Canada", ["province_or_territory_10_1"]],
        }

        # Extend mandatory fields based on additional logic
        for key, value in additional_fields.items():
            if data.get(key) == value[0]:
                mandatory_fields.extend(value[1])

        # Handle nested logic for categorizations_8 field (JSON/dict)
        category_8 = data.get("categorizations_8")
        if category_8:
            if category_8.get("isChecked"):
                if not any(
                    [
                        category_8.get("businessInCanada"),
                        category_8.get("hasAssetsInCanada"),
                        category_8.get("doesBusinessInCanada"),
                    ]
                ):
                    mandatory_fields.append("categorizations_8")
            if category_8.get("isCheckedone"):
                if not any(
                    [
                        category_8.get("largeRevenue"),
                        category_8.get("largeAssetSize"),
                        category_8.get("largeEmployees"),
                    ]
                ):
                    if "categorizations_8" not in mandatory_fields:
                        mandatory_fields.append("categorizations_8")

        missing_fields = [field for field in mandatory_fields if not data.get(field)]
        return not missing_fields

    def validate_annual_report(self, data):
        # Mandatory fields for Annual Report
        mandatory_fields_ar = [
            "steps_taken_1",
            "structure_3",
            "categorization_4",
            "policies_in_place_6",
            "risk_identified_8",
            "risk_activaties_9",
            "measures_remediate_activaties_11",
            "measures_taken_loss_income_13",
            "training_provided_15",
            "policies_procedures_assess_17",
        ]

        additional_fields_ar = {
            "steps_taken_1": [
                "Other, please specify:",
                ["steps_taken_description_1"],
            ],
            "policies_in_place_6": ["Yes", ["elements_implemented_6_1"]],
            "risk_identified_8": ["Yes", ["risk_aspects_8_1"]],
            "risk_aspects_8_1": [
                "Other, please specify:",
                ["risk_aspects_description_8_1"],
            ],
            "risk_activaties_9": [
                "Other, please specify:",
                ["risk_activaties_description_9"],
            ],
            "measures_remediate_activaties_11": [
                "Yes",
                ["remediation_measures_taken_11_1"],
            ],
            "remediation_measures_taken_11_1": [
                "Other, please specify:",
                ["remediation_measures_taken_description_11_1"],
            ],
            "training_provided_15": [
                "Yes",
                ["training_mandatory_15_1"],
            ],
            "policies_procedures_assess_17": [
                "Yes",
                ["assessment_method_17_1"],
            ],
            "assessment_method_17_1": [
                "Other, please specify:",
                ["assessment_method_description_17_1"],
            ],
        }

        for key, value in additional_fields_ar.items():
            if data.get(key) == value[0]:
                mandatory_fields_ar.extend(value[1])

        missing_fields = [field for field in mandatory_fields_ar if not data.get(field)]
        return not missing_fields

    def fill_identifying_info_sheet(self, wb, ii_data):
        """
        Populate the 'IdentifyingInformation' sheet in the workbook.
        """
        ws = wb["IdentifyingInformation"]
        ii_7_1_dict = {
            "1": "The United Kingdom’s Modern Slavery Act 2015",
            "2": "Australia’s Modern Slavery Act 2018",
            "3": "California’s Transparency in Supply Chains Act",
            "other": "",
        }
        ii_9_dict = {
            "1": "Agriculture, forestry, fishing and hunting",
            "2": "Administrative and support, waste management and remediation services",
            "3": "Mining, quarrying, and oil and gas extraction",
            "4": "Utilities",
            "5": "Construction",
            "6": "Wholesale trade",
            "7": "Retail trade",
            "8": "Transportation and warehousing",
            "9": "Information and cultural industries",
            "10": "Finance and insurance",
            "11": "Real estate and rental and leasing",
            "12": "Professional, scientific and technical services",
            "13": "Management of companies and enterprises",
            "14": "Educational services",
            "15": "Health care and social assistance",
            "16": "Arts, entertainment and recreation",
            "17": "Accommodation and food services",
            "18": "Other services (except public administration)",
            "19": "Public administration",
            "other": "",
        }

        ws["B2"] = ii_data.get("report_purpose_1")
        ws["B3"] = ii_data.get("reporting_legal_name_2")
        ws["B4"] = ii_data.get("financial_reporting_year_from_3")
        ws["B5"] = ii_data.get("financial_reporting_year_to_3")
        ws["B6"] = ii_data.get("is_revised_version_4")
        ws["B7"] = (
            ii_data.get("original_report_date_4_1")
            if ii_data.get("is_revised_version_4") == "Yes"
            else None
        )
        ws["B8"] = (
            ii_data.get("changes_description_4_2")
            if ii_data.get("is_revised_version_4") == "Yes"
            else None
        )
        ws["B9"] = ii_data.get("business_number_5")
        ws["B10"] = ii_data.get("is_joint_report_6")

        if ii_data.get("is_joint_report_6") == "Yes":
            ans_6_1_2 = "\n".join(
                [
                    f"{item.get('legalName')} : {item.get('businessNumber')}"
                    for item in ii_data.get("legal_name_and_business_numbers_6_1_2", [])
                ]
            )
        else:
            ans_6_1_2 = None
        ws["B11"] = ans_6_1_2

        ws["B12"] = ii_data.get("subject_to_supply_chain_legislation_7")
        if ii_data.get("subject_to_supply_chain_legislation_7") == "Yes":
            applicable_laws = ii_data.get("applicable_laws_7_1", [])
            ans_7_1 = ",\n".join([ii_7_1_dict.get(val, "") for val in applicable_laws])
            if "other" in applicable_laws:
                ans_7_1 += ii_data.get("other_laws_description_7_1")
        else:
            ans_7_1 = ""
        ws["B13"] = ans_7_1

        categorizations = ii_data.get("categorizations_8", {})
        ws["B15"] = "True" if categorizations.get("isChecked") else "False"
        ws["B16"] = "True" if categorizations.get("isCheckednew") else "False"
        ws["B17"] = "True" if categorizations.get("businessInCanada") else "False"
        ws["B18"] = "True" if categorizations.get("doesBusinessInCanada") else "False"
        ws["B19"] = "True" if categorizations.get("hasAssetsInCanada") else "False"
        ws["B20"] = "True" if categorizations.get("isCheckedone") else "False"
        ws["B21"] = "True" if categorizations.get("largeAssetSize") else "False"
        ws["B22"] = "True" if categorizations.get("largeRevenue") else "False"
        ws["B23"] = "True" if categorizations.get("largeEmployees") else "False"

        sectors = ii_data.get("sectors_or_industries_9", [])
        ans_9 = ",\n".join([ii_9_dict.get(val, "") for val in sectors])
        if "other" in sectors:
            ans_9 += ii_data.get("sectors_or_industries_description_9", "")
        ws["B24"] = ans_9

        ws["B25"] = ii_data.get("country_10")
        ws["B26"] = (
            ii_data.get("province_or_territory_10_1")
            if ii_data.get("country_10") == "Canada"
            else None
        )

    def fill_annual_report_sheet(self, wb, ar_data):
        """
        Populate the 'AnnualReport' sheet in the workbook.
        """
        ws = wb["AnnualReport"]

        ar_dict_6_1 = {
            "1": "Embedding responsible business conduct into policies and management systems",
            "2": "Identifying and assessing adverse impacts in operations, supply chains and business relationships",
            "3": "Ceasing, preventing or mitigating adverse impacts",
            "4": "Tracking implementation and results",
            "5": "Communicating how impacts are addressed",
            "6": "Providing for or cooperating in remediation when appropriate",
        }
        ar_dict_8 = {
            "Yes": "Yes, we have identified risks to the best of our knowledge and will continue to strive to identify emerging risks.",
            "Yesone": "Yes, we have started the process of identifying risks, but there are still gaps in our assessments.",
            "No": "No, we have not started the process of identifying risks.",
        }
        ar_dict_11 = {
            "Yes": "Yes, we have taken remediation measures and will continue to identify and address any gaps in our response.",
            "Yesone": "Yes, we have taken some remediation measures, but there are gaps in our response that still need to be addressed.",
            "No": "No, we have not taken any remediation measures.",
            "Noone": "Not applicable, we have not identified any forced labour or child labour in our activities and supply chains.",
        }
        ar_dict_13 = {
            "Yes": "Yes, we have taken substantial remediation measures and will continue to identify and address any gaps in our response.",
            "Yesone": "Yes, we have taken some remediation measures, but there are gaps in our response that still need to be addressed.",
            "No": "No, we have not taken any remediation measures.",
            "Noone": "Not applicable, we have not identified any loss of income to vulnerable families resulting from measures taken to eliminate the use of forced labour or child labour in our activities and supply chains.",
        }
        ar_dict_15_1 = {
            "Yes": "Yes, the training is mandatory for all employees.",
            "Yesone": "Yes, the training is mandatory for employees making contracting or purchasing decisions.",
            "Yestwo": "Yes, the training is mandatory for some employees.",
            "No": "No, the training is voluntary.",
        }

        # Process steps_taken_1
        steps = ar_data.get("steps_taken_1", [])
        if "Other, please specify:" in steps:
            steps_filtered = [i for i in steps if i != "Other, please specify:"]

            if steps_filtered:
                ws["B2"] = ",\n".join(steps_filtered) + (
                    ",\n" + ar_data.get("steps_taken_description_1", "")
                    if ar_data.get("steps_taken_description_1")
                    else ""
                )
            else:
                ws["B2"] = ar_data.get("steps_taken_description_1", "")
        else:
            ws["B2"] = ",\n".join(steps)

        ws["B3"] = ar_data.get("additional_information_2")
        ws["B4"] = ar_data.get("structure_3")

        categorization = ar_data.get("categorization_4", {})
        ws["B6"] = "True" if categorization.get("Producing_goods") else "False"
        ws["B7"] = "True" if categorization.get("ProducinggoodsInCanada") else "False"
        ws["B8"] = (
            "True" if categorization.get("ProducinggoodsOutsideCanada") else "False"
        )
        ws["B9"] = "True" if categorization.get("Selling_goods") else "False"
        ws["B10"] = "True" if categorization.get("SellinggoodsInCanada") else "False"
        ws["B11"] = (
            "True" if categorization.get("SellinggoodsOutsideCanada") else "False"
        )
        ws["B12"] = "True" if categorization.get("Distributing_goods") else "False"
        ws["B13"] = (
            "True" if categorization.get("DistributinggoodsInCanada") else "False"
        )
        ws["B14"] = (
            "True" if categorization.get("DistributinggoodsOutsideCanada") else "False"
        )
        ws["B15"] = (
            "True"
            if categorization.get("Importing_into_Canada_goods_produced_outside_Canada")
            else "False"
        )
        ws["B16"] = (
            "True"
            if categorization.get("Controlling_an_entity_engaged_in_producing")
            else "False"
        )

        ws["B17"] = ar_data.get("additional_information_entity_5")
        ws["B18"] = ar_data.get("policies_in_place_6")
        elements = ar_data.get("elements_implemented_6_1", [])
        ws["B19"] = ",\n".join([ar_dict_6_1.get(item, "") for item in elements])
        ws["B20"] = ar_data.get("additional_info_policies_7")
        ws["B21"] = ar_dict_8.get(ar_data.get("risk_identified_8"), "")

        risk_aspects = ar_data.get("risk_aspects_8_1", [])
        if "Other, please specify:" in risk_aspects:
            aspects_filtered = [
                i for i in risk_aspects if i != "Other, please specify:"
            ]
            if aspects_filtered:
                ws["B22"] = ",\n".join(aspects_filtered) + (
                    ",\n" + ar_data.get("risk_aspects_description_8_1", "")
                    if ar_data.get("risk_aspects_description_8_1")
                    else ""
                )
            else:
                ws["B22"] = ar_data.get("risk_aspects_description_8_1", "")
        else:
            ws["B22"] = ",\n".join(risk_aspects)

        risk_activities = ar_data.get("risk_activaties_9", [])
        if "Other, please specify:" in risk_activities:
            activities_filtered = [
                i for i in risk_activities if i != "Other, please specify:"
            ]
            if activities_filtered:
                ws["B23"] = ",\n".join(activities_filtered) + (
                    ",\n" + ar_data.get("risk_activaties_description_9", "")
                    if ar_data.get("risk_activaties_description_9")
                    else ""
                )
            else:
                ws["B23"] = ar_data.get("risk_activaties_description_9", "")
        else:
            ws["B23"] = ",\n".join(risk_activities)

        ws["B24"] = ar_data.get("additional_info_entity_10")
        ws["B25"] = ar_dict_11.get(ar_data.get("measures_remediate_activaties_11"), "")

        remediation = ar_data.get("remediation_measures_taken_11_1", [])
        if "Other, please specify:" in remediation:
            remediation_filtered = [
                i for i in remediation if i != "Other, please specify:"
            ]

            if remediation_filtered:
                ws["B26"] = ",\n".join(remediation_filtered) + (
                    ",\n"
                    + ar_data.get("remediation_measures_taken_description_11_1", "")
                    if ar_data.get("remediation_measures_taken_description_11_1")
                    else ""
                )
            else:
                ws["B26"] = ar_data.get(
                    "remediation_measures_taken_description_11_1", ""
                )

        else:
            ws["B26"] = ",\n".join(remediation)

        ws["B27"] = ar_data.get("remediation_measures_12")
        ws["B28"] = ar_dict_13.get(ar_data.get("measures_taken_loss_income_13"), "")
        ws["B29"] = ar_data.get("additional_info_loss_income_14")
        ws["B30"] = ar_data.get("training_provided_15")
        if ar_data.get("training_provided_15") == "Yes":
            ws["B31"] = ar_dict_15_1.get(ar_data.get("training_mandatory_15_1"), "")
        else:
            ws["B31"] = None
        ws["B32"] = ar_data.get("additional_info_training_16")
        ws["B33"] = ar_data.get("policies_procedures_assess_17")
        if ar_data.get("policies_procedures_assess_17") == "Yes":
            assessment = ar_data.get("assessment_method_17_1", [])
            if "Other, please specify:" in assessment:
                assess_filtered = [
                    i for i in assessment if i != "Other, please specify:"
                ]
                if assess_filtered:
                    ws["B34"] = ",\n".join(assess_filtered) + (
                        ",\n" + ar_data.get("assessment_method_description_17_1", "")
                        if ar_data.get("assessment_method_description_17_1")
                        else ""
                    )
                else:
                    ws["B34"] = ar_data.get("assessment_method_description_17_1", "")
            else:
                ws["B34"] = ",\n".join(assessment)
        else:
            ws["B34"] = None
        ws["B35"] = ar_data.get("additional_info_assessment_18")

    def get(self, request):
        try:
            # Validate and extract query parameters
            serializer = CheckOrgCoprSerializer(data=request.query_params)
            serializer.is_valid(raise_exception=True)
            organization = serializer.validated_data["organization"]
            corporate = serializer.validated_data.get("corporate")
            year = int(request.query_params["year"])

            ii_exist = True
            ar_exist = True
            ii = None
            ar = None

            # Retrieve Identifying Information and Annual Report objects based on corporate presence
            if corporate:
                try:
                    ii = IdentifyingInformation.objects.get(
                        corporate__id=corporate.id,
                        organization__id=organization.id,
                        year=year,
                    )
                except IdentifyingInformation.DoesNotExist:
                    ii_exist = False

                try:
                    ar = AnnualReport.objects.get(
                        corporate__id=corporate.id,
                        organization__id=organization.id,
                        year=year,
                    )
                except AnnualReport.DoesNotExist:
                    ar_exist = False
            else:
                try:
                    ii = IdentifyingInformation.objects.get(
                        organization__id=organization.id,
                        year=year,
                        corporate__isnull=True,
                    )
                except IdentifyingInformation.DoesNotExist:
                    ii_exist = False

                try:
                    ar = AnnualReport.objects.get(
                        organization__id=organization.id,
                        year=year,
                        corporate__isnull=True,
                    )
                except AnnualReport.DoesNotExist:
                    ar_exist = False

            # If neither exists, return an error response
            if not ii_exist and not ar_exist:
                return Response(
                    {
                        "message": "Please fill the mandatory details for Identifying Information and Annual Report"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate data (if object exists)
            ii_valid = True
            ar_valid = True

            if ii_exist:
                ii_data_dict = model_to_dict(ii)
                if not self.validate_identifying_info(ii_data_dict):
                    ii_valid = False
                else:
                    ii_serializer = IISerializer(data=ii_data_dict, screen_number=None)
                    ii_serializer.is_valid(raise_exception=True)
                    ii_data = ii_serializer.validated_data

            if ar_exist:
                ar_data_dict = model_to_dict(ar)
                if not self.validate_annual_report(ar_data_dict):
                    ar_valid = False
                else:
                    ar_serializer = ARSerializer(data=ar_data_dict, screen_number=None)
                    ar_serializer.is_valid(raise_exception=True)
                    ar_data = ar_serializer.validated_data

            # Return error if any existing data is invalid
            if ii_exist and not ii_valid and ar_exist and not ar_valid:
                return Response(
                    {
                        "message": "Please fill the mandatory details for Identifying Information and Annual Report"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            elif ii_exist and not ii_valid:
                return Response(
                    {
                        "message": "Please fill the mandatory details for Identifying Information"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            elif ar_exist and not ar_valid:
                return Response(
                    {"message": "Please fill the mandatory details for Annual Report"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Load the template and fill in the sheets
                    
            blob_name = "CanadaBillS211.xlsx"
            blob_object = default_storage.open(blob_name, "rb")
            blob_stream = BytesIO(blob_object.read())


            # template_path = os.path.join(
            #     settings.BASE_DIR, "templates", "CanadaBillS211", "CanadaBillS211.xlsx"
            # )
            wb = load_workbook(blob_stream)

            if ii_exist and ii_valid:
                self.fill_identifying_info_sheet(wb, ii_data)
            if ar_exist and ar_valid:
                self.fill_annual_report_sheet(wb, ar_data)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                "attachment; filename=Canada_Bill_S211.xlsx"
            )
            return response

        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            tb_details = traceback.extract_tb(exc_traceback)
            error_info = {
                "error": str(e),
                "exception_type": str(exc_type.__name__),
                "file": tb_details[-1].filename,
                "line": tb_details[-1].lineno,
                "function": tb_details[-1].name,
                "code": tb_details[-1].line,
            }
            return Response(error_info, status=status.HTTP_400_BAD_REQUEST)
