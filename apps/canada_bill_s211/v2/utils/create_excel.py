import openpyxl
from django.core.files.storage import default_storage
from django.contrib.auth import get_user_model
from sustainapp.models import Organization, Corporateentity
from apps.canada_bill_s211.v2.models.ReportingForEntities import ReportingForEntities
from apps.canada_bill_s211.v2.models.SubmissionInformation import SubmissionInformation
from bs4 import BeautifulSoup
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from apps.canada_bill_s211.v2.constants import REPORTING_FOR_ENTITIES_RADIO_BUTTON
import logging
import io

logger = logging.getLogger("django")
CustomUser = get_user_model()

class CanadaBillReport:
    """
    This service file gets the data from the models on the basis of
    organization
    corporate
    year

    and then downloads the excel sheet from azure data storage
    after that edits the downloaded excel sheet as per requirements.
    """
    def __init__(self, user ,organization, corporate, year:int):
        self.excel_file_url = default_storage.open("canada_bill_s211/BillS211Template.xlsx")
        self.excel_file = openpyxl.load_workbook(self.excel_file_url)
        self.user = user
        self.organization: Organization = organization
        self.corporate: Corporateentity | None = corporate
        self.year:int = year
        self.not_available = "NA"

    def get_part_one_data(self):
        queryset = (
            SubmissionInformation.objects.filter(
            organization__client = self.user.client,
            organization__in = self.user.orgs.all(),
        ).filter(
            organization = self.organization,
            corporate = self.corporate,
            year = self.year
        )
        )
        if self.corporate is not None:
            queryset = queryset.filter(corporate__in = self.user.corps.all())
        data = dict()
        for i in queryset.order_by("screen").values_list("data",flat=True):
            data.update(i)
        return data

    def get_part_two_data(self):
        queryset = (
            ReportingForEntities.objects.filter(
            organization__client = self.user.client,
            organization__in = self.user.orgs.all(),
        ).filter(
            organization = self.organization,
            corporate = self.corporate,
            year = self.year
        )
        )
        if self.corporate is not None:
            queryset = queryset.filter(corporate__in = self.user.corps.all())
        data = dict()
        for i in queryset.order_by("screen").values_list("data",flat=True):
            data.update(i)
        return data

    def retrieve_key2_if_key1_matches(self, dictionary:dict ,key1:str, key2:str, value:str):
        if dictionary.get(key1) == value:
            return dictionary.get(key2)
        return self.not_available

    def create_and_merge_rows(self, sheet:Worksheet, row_insert_number: int, insert_data:list)-> int:
        if not insert_data:
            insert_data = [self.not_available]
        target_column = 2
        numbers_of_rows_to_insert = len(insert_data)
        # Get question cell

        answer_cell = sheet.cell(row=row_insert_number, column=target_column + 1)
        answer_cell.value = insert_data[0]
        answer_cell.alignment = Alignment(wrap_text=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for idx, data in enumerate(insert_data[1:]):
            current_row_index = answer_cell.row + idx + 1
            sheet.insert_rows(idx=current_row_index, amount=1)
            sheet.cell(row=current_row_index, column=target_column + 1 ,value=data)
            sheet.cell(row=current_row_index, column=target_column + 1).border = thin_border
            sheet.cell(row=current_row_index, column=target_column + 1).alignment = Alignment(wrap_text=True)

        sheet.merge_cells(start_row=answer_cell.row, start_column=target_column, end_row = answer_cell.row + numbers_of_rows_to_insert - 1, end_column=target_column)
        return answer_cell.row + numbers_of_rows_to_insert


    def add_in_list_if_not_null(self, original_list, probable_null_item):
        if probable_null_item is not None and original_list != [self.not_available]:
            original_list.append(probable_null_item)
        return original_list

    def modify_text_as_per_requirements(self,heading:str,points:list):
        paragraph = "\n".join(points)
        cell_text = CellRichText(
            TextBlock(InlineFont(b=True), f"{heading}\n"),
            TextBlock(InlineFont(), paragraph)
        )
        return cell_text

    def modify_recursive_list_data(self,data,ignore_keys=None):
        if ignore_keys is None:
            ignore_keys = []
        temporary_list=[]
        for i in data:
            if i not in ignore_keys:
                heading = i.replace('(select all that apply)', '').strip()
                temporary_list.append(self.modify_text_as_per_requirements(heading=heading, points=data[i]))
        return temporary_list

    def modify_part_one(self):
        """
        Modifies the Part 1 sheet.
        Gets all the data from the ReportingForEntities Model based on the user,
        organization, corporate entity, and year.
        Then based on the screen we fill the data.
        """
        #* Get data from the model
        part_one_data = self.get_part_one_data()
        part_one_sheet = self.excel_file["Submission Information"]
        part_one_sheet.cell(row=3, column=3).value = part_one_data.get("screen1_q1",self.not_available)
        part_one_sheet.cell(row=4, column=3).value = part_one_data.get("screen1_q2",self.not_available)
        part_one_sheet.cell(row=5, column=3).value = part_one_data.get("screen1_q3",self.not_available)
        part_one_sheet.cell(row=6, column=3).value = f"{part_one_data.get('screen1_to_q4',self.not_available)} - {part_one_data.get('screen1_form_q4',self.not_available)}"
        part_one_sheet.cell(row=7, column=3).value = part_one_data.get("screen2_q1",self.not_available)
        part_one_sheet.cell(row=8, column=3).value = self.retrieve_key2_if_key1_matches(dictionary=part_one_data, key1="screen2_q1", key2="screen2_q2", value="Yes")
        part_one_sheet.cell(row=9, column=3).value = self.html_to_plain_text(html_string=self.retrieve_key2_if_key1_matches(dictionary=part_one_data, key1="screen2_q1", key2="screen2_q3", value="Yes"))
        part_one_sheet.cell(row=9, column=3).alignment = Alignment(wrap_text=True)
        part_one_sheet.cell(row=10, column=3).value = part_one_data.get("screen2_q4",self.not_available)
        part_one_sheet.cell(row=11, column=3).value = part_one_data.get("screen3_q1",self.not_available)
        current_row = self.create_and_merge_rows(sheet=part_one_sheet, row_insert_number=12, insert_data=[i["legalName"] for i in part_one_data.get("screen3_q2", [self.not_available])])
        current_row = self.create_and_merge_rows(sheet=part_one_sheet, row_insert_number=current_row, insert_data=[i["businessNumber"] for i in part_one_data.get("screen3_q2", [self.not_available])])
        part_one_sheet.cell(row=current_row, column=3).value = part_one_data.get("screen4_q1",self.not_available)
        current_row+=1
        current_row = self.create_and_merge_rows(sheet=part_one_sheet, row_insert_number=current_row, insert_data=self.add_in_list_if_not_null(original_list=part_one_data.get("screen4_q2", [self.not_available]), probable_null_item=part_one_data.get("screen4_q3_other")))
        current_row = self.create_and_merge_rows(sheet=part_one_sheet, row_insert_number=current_row, insert_data=self.modify_recursive_list_data(part_one_data.get("screen5_q1")))
        current_row = self.create_and_merge_rows(sheet=part_one_sheet, row_insert_number=current_row, insert_data=self.modify_recursive_list_data(data=part_one_data.get("screen6_q1"),ignore_keys=["other"]))
        part_one_sheet.cell(row=current_row, column=3).value = part_one_data.get("screen6_q2", self.not_available)
        current_row+=1
        part_one_sheet.cell(row=current_row, column=3).value = part_one_data.get("screen7_q1", self.not_available)
        part_one_sheet.cell(row=current_row, column=3).value = part_one_data.get("screen7_q2", self.not_available)

    def modify_part_two(self):
        """
        Modifies the Part 2 sheet.
        Gets all the data from the ReportingForEntities Model based on the user,
        organization, corporate entity, and year.
        Then based on the screen we fill the data.
        """
        part_two_data = self.get_part_two_data()
        part_two_sheet = self.excel_file["Reporting For Entities"]
        part_two_sheet.cell(row=3,column=3).value = part_two_data.get("screen1_q1",self.not_available)
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=4, insert_data=part_two_data.get("screen1_q2",[self.not_available]))
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=part_two_data.get("screen2_q1",[self.not_available]))
        part_two_sheet.cell(row=current_row, column=3).value = part_two_data.get("screen2_q2",self.not_available)
        current_row+=1
        part_two_sheet.cell(row=current_row, column=3).value = part_two_data.get("screen3_q1",self.not_available)
        current_row+=1
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=part_two_data.get("screen3_q2",[self.not_available]))
        part_two_sheet.cell(row=current_row, column=3).value = REPORTING_FOR_ENTITIES_RADIO_BUTTON[4].get((part_two_data.get("screen4_q1",self.not_available)),self.not_available)
        part_two_sheet.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)
        current_row+=1
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=part_two_data.get("screen4_q2",[self.not_available]))
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=self.add_in_list_if_not_null(original_list=self.modify_recursive_list_data(data=part_two_data.get("screen5_q1"),ignore_keys="other"),probable_null_item=part_two_data.get("screen5_q2",[self.not_available])))
        part_two_sheet.cell(row=current_row, column=3).value = part_two_data.get("screen5_q3",self.not_available)
        current_row+=1
        part_two_sheet.cell(row=current_row, column=3).value = REPORTING_FOR_ENTITIES_RADIO_BUTTON[6].get((part_two_data.get("screen6_q1",self.not_available)),self.not_available)
        part_two_sheet.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)
        current_row+=1
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=part_two_data.get("screen6_q2",[self.not_available]))
        part_two_sheet.cell(row=current_row, column=3).value = REPORTING_FOR_ENTITIES_RADIO_BUTTON[7][0].get((part_two_data.get("screen7_q1",self.not_available)),self.not_available)
        part_two_sheet.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)
        current_row+=1
        part_two_sheet.cell(row=current_row, column=3).value = REPORTING_FOR_ENTITIES_RADIO_BUTTON[7][1].get((part_two_data.get("screen7_q2",self.not_available)),self.not_available)
        part_two_sheet.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)
        current_row+=1
        part_two_sheet.cell(row=current_row, column=3).value = REPORTING_FOR_ENTITIES_RADIO_BUTTON[7][2].get((part_two_data.get("screen7_q3",self.not_available)),self.not_available)
        part_two_sheet.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)
        current_row+=1
        part_two_sheet.cell(row=current_row, column=3).value = part_two_data.get("screen8_q1",self.not_available)
        current_row+=1
        current_row = self.create_and_merge_rows(sheet=part_two_sheet, row_insert_number=current_row, insert_data=part_two_data.get("screen8_q2",[self.not_available]))
        part_two_sheet.cell(row=current_row, column=3).value = part_two_data.get("screen8_q3",self.not_available)


    def html_to_plain_text(self,html_string):
        """
        Converts an HTML string to plain text, attempting to preserve basic formatting.
        """
        if not isinstance(html_string, str): # Ensure input is a string
            return str(html_string) if html_string is not None else ""

        soup = BeautifulSoup(html_string, 'html.parser')
        for tag in soup(['p', 'br', 'div', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6']):
            tag.append('\n')
        plain_text = soup.get_text()
        plain_text = '\n'.join([line.strip() for line in plain_text.splitlines() if line.strip()])
        return plain_text

    def insert_rows_in_sheet(self, sheet_name: str, start_row: int, num_rows_to_insert: int):
        """
        Inserts a specified number of empty rows into a given sheet.

        Args:
            sheet_name: The name of the Excel sheet to modify.
            start_row: The row number (1-indexed) before which new rows will be inserted.
            num_rows_to_insert: The number of empty rows to insert.
        """
        if num_rows_to_insert <= 0:
            logger.info(f"No rows to insert requested for sheet '{sheet_name}'.")
            return

        try:
            sheet = self.excel_file[sheet_name]
            # Insert rows one by one to preserve formatting if openpyxl's insert_rows has issues with it
            # However, openpyxl's insert_rows should generally handle this correctly.
            # For complex scenarios with merged cells or specific styles, more granular control might be needed.
            sheet.insert_rows(idx=start_row, amount=num_rows_to_insert)
            logger.info(f"Inserted {num_rows_to_insert} rows into sheet '{sheet_name}' starting before row {start_row}.")
        except KeyError:
            logger.error(f"Sheet '{sheet_name}' not found in the Excel file.")
        except Exception as e:
            logger.error(f"Error inserting rows into sheet '{sheet_name}': {e}")

    def generate_excel_report_data(self):
        """
        Modifies the Excel sheet and returns its content as bytes.
        """
        self.modify_part_one()
        self.modify_part_two()
        # Example of how you might use insert_rows_in_sheet:
        # self.insert_rows_in_sheet(sheet_name="Part 1", start_row=12, num_rows_to_insert=5) # Adjust start_row as needed

        # Add calls to other modification methods here if they exist (e.g., self.modify_part_two())

        excel_io = io.BytesIO()
        self.excel_file.save(excel_io)
        excel_io.seek(0) # Go to the beginning of the stream
        return excel_io.getvalue()
