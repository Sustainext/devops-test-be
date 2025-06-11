from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import base64
from django.db import transaction
from datametric.models import RawResponse, Path
from sustainapp.models import Location
import uuid
from django.core.cache import cache


class ExcelTemplateUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    expected_headers = [
        "Location",
        "Month",
        "Year",
        "Scope",
        "Category",
        "Subcategory",
        "Activity",
        "Quantity 1",
        "Unit 1",
        "Quantity 2",
        "Unit 2",
    ]

    required_fields = {
        "Location",
        "Month",
        "Year",
        "Scope",
        "Category",
        "Subcategory",
        "Quantity 1",
    }

    max_file_size = 5 * 1024 * 1024  # 5 MB

    VALID_SCOPES = {"Scope 1", "Scope 2", "Scope 3"}

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Combined scope data from your input
        self.scope_data = {
            "Scope 1": [
                {"name": "Stationary Combustion", "SubCategory": ["Fuel"]},
                {
                    "name": "Mobile Combustion",
                    "SubCategory": ["Fuel", "Vehicles", "Road Travel", "Rail Freight"],
                },
                {
                    "name": "Refrigerants and Fugitive Gases",
                    "SubCategory": ["Refrigerants and Fugitive Gases"],
                },
            ],
            "Scope 2": [
                {
                    "name": "Purchased Electricity",
                    "SubCategory": ["Electricity", "Energy Services", "Utilities"],
                },
                {
                    "name": "Purchased Heat & Steam",
                    "SubCategory": ["Heat and Steam", "Energy Services", "Utilities"],
                },
                {
                    "name": "Purchased Cooling",
                    "SubCategory": ["Electricity", "Energy Services", "Utilities"],
                },
            ],
            "Scope 3": [
                {
                    "name": "Purchased Goods & Services",
                    "SubCategory": [
                        "Agriculture/Hunting/Forestry/Fishing",
                        "Arable Farming",
                        "Building Materials",
                        "Ceramic Goods",
                        "Chemical Products",
                        "Clothing and Footwear",
                        "Cloud Computing - CPU",
                        "Cloud Computing - Memory",
                        "Cloud Computing - Networking",
                        "Cloud Computing - Storage",
                        "Consumer Goods Rental",
                        "DIY and Gardening Equipment",
                        "Domestic Services",
                        "Education",
                        "Electrical Equipment",
                        "Electronics",
                        "Equipment Rental",
                        "Fabricated Metal Products",
                        "Financial Services",
                        "Fishing/Aquaculture/Hunting",
                        "Food and Beverage Services",
                        "Food/Beverages/Tobacco",
                        "Furnishings and Household",
                        "General Retail",
                        "Glass and Glass Products",
                        "Government Activities",
                        "Health and Social Care",
                        "Health Care",
                        "Housing",
                        "Information and Communication Services",
                        "Insurance Services",
                        "Livestock Farming",
                        "Machinery",
                        "Maintenance and Repair",
                        "Manufacturing",
                        "Metals",
                        "Mined Materials",
                        "Mining",
                        "Non-profit Activities",
                        "Office Equipment",
                        "Operational Activities",
                        "Organic Products",
                        "Other Materials",
                        "Paper and Cardboard",
                        "Paper Products",
                        "Pavement and Surfacing",
                        "Personal Care and Accessories",
                        "Plastics and Rubber Products",
                        "Professional Services and Activities",
                        "Real Estate",
                        "Recreation and Culture",
                        "Social Care",
                        "Textiles",
                        "Timber and Forestry Products",
                        "Utilities",
                        "Vehicle Maintenance and Services",
                        "Vehicle Parts",
                        "Waste Management",
                        "Water Supply",
                        "Water Treatment",
                        "Wholesale Trade",
                    ],
                },
                {
                    "name": "Capital Goods",
                    "SubCategory": [
                        "Construction",
                        "Electrical Equipment",
                        "Electronics",
                        "Furnishings and Household",
                        "Health Care",
                        "Machinery",
                        "Office Equipment",
                    ],
                },
                {
                    "name": "Fuel & Energy Related Activities",
                    "SubCategory": ["Electricity", "Fuel"],
                },
                {
                    "name": "Upstream Transportation & Distribution",
                    "SubCategory": [
                        "Air Freight",
                        "Energy Services",
                        "Rail Freight",
                        "Road Freight",
                        "Sea Freight",
                        "Transport Services and Warehousing",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Waste Generated in Operations",
                    "SubCategory": [
                        "Construction Waste",
                        "Electrical Waste",
                        "Food and Organic Waste",
                        "General Waste",
                        "Glass Waste",
                        "Metal Waste",
                        "Paper and Cardboard Waste",
                        "Plastic Waste",
                        "Waste Management",
                        "Water Treatment",
                    ],
                },
                {
                    "name": "Business Travel",
                    "SubCategory": [
                        "Restaurants and Accommodation",
                        "Accommodation",
                        "Road Travel",
                        "Air Travel",
                        "Rail Travel",
                        "Sea Travel",
                        "Tickets and Passes",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Employee Commuting",
                    "SubCategory": [
                        "Homeworking",
                        "Rail Travel",
                        "Road Travel",
                        "Tickets and Passes",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Upstream Leased Assets",
                    "SubCategory": ["Facility", "Housing", "Real Estate"],
                },
                {
                    "name": "Downstream Transportation & Distribution",
                    "SubCategory": [
                        "Air Freight",
                        "Energy Services",
                        "Infrastructure",
                        "Rail Freight",
                        "Road Freight",
                        "Sea Freight",
                        "Transport Services and Warehousing",
                    ],
                },
                {
                    "name": "Processing of Sold Products",
                    "SubCategory": [
                        "Cloud Computing - CPU",
                        "Cloud Computing - Memory",
                        "Cloud Computing - Networking",
                        "Cloud Computing - Storage",
                        "Information and Communication Services",
                    ],
                },
                {
                    "name": "End of Life Treatment of Sold Products",
                    "SubCategory": ["Electricity", "Fuel"],
                },
                {
                    "name": "Downstream Leased Assets",
                    "SubCategory": [
                        "Equipment Rental",
                        "Facility",
                        "Housing",
                        "Real Estate",
                    ],
                },
            ],
        }

        self.units = [
            "m2",
            "km2",
            "ft2",
            "ha",
            "containers",
            "MB",
            "GB",
            "TB",
            "m",
            "km",
            "ft",
            "mi",
            "nmi",
            "Wh",
            "kWh",
            "MWh",
            "MJ",
            "GJ",
            "TJ",
            "BTU",
            "therm",
            "MMBTU",
            "usd",
            "afn",
            "dzd",
            "ars",
            "aud",
            "bhd",
            "brl",
            "cad",
            "kyd",
            "cny",
            "dkk",
            "egp",
            "eur",
            "hkd",
            "huf",
            "isk",
            "inr",
            "iqd",
            "ils",
            "jpy",
            "lbp",
            "mxn",
            "mad",
            "nzd",
            "nok",
            "qar",
            "rub",
            "sar",
            "sgd",
            "zar",
            "krw",
            "sek",
            "chf",
            "thb",
            "twd",
            "tnd",
            "try",
            "aed",
            "gbp",
            "number of Nights",
            "numbers",
            "passengers",
            "ms",
            "s",
            "m",
            "h",
            "day",
            "year",
            "ml",
            "l",
            "m3",
            "standard_cubic_foot",
            "gallon_us",
            "bbl",
            "g",
            "kg",
            "t",
            "ton",
            "lb",
        ]

        self.units2 = ["ms", "s", "m", "h", "day", "year", "m", "km", "ft", "mi", "nmi"]

        self.months = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]

        # Build flattened mappings
        self.scope_to_categories = {}
        self.category_to_subcategories = {}

        for scope, categories in self.scope_data.items():
            cat_names = []
            for cat in categories:
                cat_name = cat["name"]
                cat_names.append(cat_name)
                self.category_to_subcategories[cat_name] = cat["SubCategory"]
            self.scope_to_categories[scope] = cat_names

    def validate_data(self, worksheet, request):
        valid_rows = []
        invalid_rows = []
        valid_locations = request.user.locs.all()

        for idx, row in enumerate(
            worksheet.iter_rows(
                min_row=2, max_col=len(self.expected_headers), values_only=True
            ),
            start=2,
        ):
            if not any(row):
                continue

            row_dict = dict(zip(self.expected_headers, row))
            missing = [
                field for field in self.required_fields if not row_dict.get(field)
            ]
            dropdown_errors = []

            # Dropdown validations
            location = row_dict.get("Location")
            scope = row_dict.get("Scope")
            month = row_dict.get("Month")
            year = str(row_dict.get("Year")) if row_dict.get("Year") else None
            category = row_dict.get("Category")
            subcategory = row_dict.get("Subcategory")
            unit1 = row_dict.get("Unit 1")
            unit2 = row_dict.get("Unit 2")

            if location and location not in [loc.name for loc in valid_locations]:
                dropdown_errors.append(f"Invalid Location: {location}")

            if scope and scope not in self.VALID_SCOPES:
                dropdown_errors.append(f"Invalid Scope: {scope}")

            if month and month not in self.months:
                dropdown_errors.append(f"Invalid Month: {month}")

            if year and year not in {str(y) for y in range(2018, 2031)}:
                dropdown_errors.append(f"Invalid Year: {year}")

            if scope in self.scope_to_categories:
                if category and category not in self.scope_to_categories[scope]:
                    dropdown_errors.append(
                        f"Category '{category}' not allowed under Scope '{scope}'"
                    )
            else:
                if category:
                    dropdown_errors.append(
                        f"Invalid Scope '{scope}' for Category check"
                    )

            if category in self.category_to_subcategories:
                if (
                    subcategory
                    and subcategory not in self.category_to_subcategories[category]
                ):
                    dropdown_errors.append(
                        f"Subcategory '{subcategory}' not allowed under Category '{category}'"
                    )
            else:
                if subcategory:
                    dropdown_errors.append(
                        f"Invalid Category '{category}' for Subcategory check"
                    )

            # Optional: Validate units
            if unit1 and unit1 not in self.units:
                dropdown_errors.append(f"Invalid Unit 1: {unit1}")

            if unit2 and unit2 not in self.units2:
                dropdown_errors.append(f"Invalid Unit 2: {unit2}")

            if missing or dropdown_errors:
                row_dict["Error"] = ""
                if missing:
                    row_dict["Error"] += f"Missing: {', '.join(missing)}. "
                if dropdown_errors:
                    row_dict["Error"] += f"Invalid: {', '.join(dropdown_errors)}"
                row_dict["Row"] = idx
                invalid_rows.append(row_dict)
            else:
                valid_rows.append(row_dict)

        return valid_rows, invalid_rows

    def save_valid_rows_to_raw_response(self, request, valid_rows):
        """
        Save valid rows to RawResponse grouped by (scope, location, year, month).
        Always append rows without checking for duplicates.
        """

        SCOPE_TO_PATH = {
            "Scope 1": "gri-environment-emissions-301-a-scope-1",
            "Scope 2": "gri-environment-emissions-301-a-scope-2",
            "Scope 3": "gri-environment-emissions-301-a-scope-3",
        }

        month_mapping = {
            "Jan": 1,
            "Feb": 2,
            "Mar": 3,
            "Apr": 4,
            "May": 5,
            "Jun": 6,
            "Jul": 7,
            "Aug": 8,
            "Sep": 9,
            "Oct": 10,
            "Nov": 11,
            "Dec": 12,
        }

        grouped_data = {}
        for row in valid_rows:
            scope = row["Scope"]
            location = row["Location"]
            year = row["Year"]
            month = row["Month"]
            key = (scope, location, year, month)
            grouped_data.setdefault(key, []).append(row)

        with transaction.atomic():
            for (scope, location, year, month), rows in grouped_data.items():
                path_slug = SCOPE_TO_PATH.get(scope)
                if not path_slug:
                    continue

                try:
                    path = Path.objects.get(slug=path_slug)
                except Path.DoesNotExist:
                    continue

                try:
                    locale_obj = Location.objects.get(name=location)
                except Location.DoesNotExist:
                    continue  # skip invalid location

                raw_response, _ = RawResponse.objects.get_or_create(
                    path=path,
                    client=request.user.client,
                    locale=locale_obj,
                    year=year,
                    month=month_mapping.get(month, 0),
                    user=request.user,
                    defaults={"data": []},
                )

                existing_data = (
                    raw_response.data if isinstance(raw_response.data, list) else []
                )

                for row in rows:
                    emission = {
                        "assigned_to": "",
                        "Category": row["Category"],
                        "Subcategory": row["Subcategory"],
                        "Activity": row["Activity"],
                        "activity_id": "",
                        "unit_type": "",
                        "Unit": row.get("Unit 1", ""),
                        "Quantity": row.get("Quantity 1", ""),
                        "Unit2": row.get("Unit 2", ""),
                        "Quantity2": row.get("Quantity 2", ""),
                        "file": {
                            "name": "",
                            "url": "",
                            "type": "",
                            "size": "",
                            "uploadDateTime": "",
                            "uploadedBy": "",
                        },
                        "rowType": "bulk_upload",
                        "scope": scope.lower().replace(" ", "_"),
                    }

                    existing_data.append({"Emission": emission, "id": None})

                raw_response.data = existing_data
                raw_response.save()

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response(
                {"message": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST
            )

        if not excel_file.name.endswith(".xlsx"):
            return Response(
                {"message": "Only .xlsx files are allowed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if excel_file.size > self.max_file_size:
            return Response(
                {"message": "File size exceeds 5MB limit."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            file_content = excel_file.read()
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        except Exception as e:
            return Response(
                {"message": f"Invalid Excel file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if "Template" not in wb.sheetnames:
            return Response(
                {"message": "Missing 'Template' sheet."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb["Template"]
        header_row = [cell.value for cell in ws[1]]
        if header_row != self.expected_headers:
            return Response(
                {"message": "Template headers do not match expected format."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        valid_rows, invalid_rows = self.validate_data(ws, request)
        error_file = self._create_error_excel(invalid_rows) if invalid_rows else None
        row_status = None
        if len(invalid_rows) > 0:
            row_status = 400
        else:
            row_status = 200

        # Store file and validation results in Redis
        temp_id = str(uuid.uuid4())
        cache_key = f"excel_upload:{temp_id}"
        cache.set(
            cache_key,
            {
                "file_content": file_content,
                "user_id": request.user.id,
            },
            timeout=900,
        )  # expires in 15mins

        return Response(
            {
                "message": "Validation complete. Use confirm API to proceed.",
                "valid_count": len(valid_rows),
                "invalid_count": len(invalid_rows),
                "valid_preview": valid_rows[:5],
                "invalid_preview": invalid_rows[:5],
                "temp_id": temp_id,
                "error_file_base64": error_file,
                "row_status": row_status,
            },
            status=status.HTTP_200_OK,
        )

    def _create_error_excel(self, error_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invalid Rows"

        # Headers
        headers = self.expected_headers + ["Error", "Row"]
        ws.append(headers)

        for row in error_rows:
            row_data = [row.get(h) for h in headers]
            ws.append(row_data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode("utf-8")
