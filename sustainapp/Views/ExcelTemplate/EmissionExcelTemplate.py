# views.py

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
import io
import re
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from openpyxl.styles import PatternFill, Font, Border, Side


class ExcelTemplateDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Combined scope data from your input
        self.scope_data = {
            "Scope 1": [
                {"name": "Stationary Combustion", "SubCategory": ["Fuel"]},
                {
                    "name": "Mobile Combustion",
                    "SubCategory": ["Fuel", "Vehicles", "Road Travel", "Rail Freight"],
                },
                {
                    "name": "Refrigerants and Fugitive Gases",
                    "SubCategory": ["Refrigerants and Fugitive Gases"],
                },
            ],
            "Scope 2": [
                {
                    "name": "Purchased Electricity",
                    "SubCategory": ["Electricity", "Energy Services", "Utilities"],
                },
                {
                    "name": "Purchased Heat & Steam",
                    "SubCategory": ["Heat and Steam", "Energy Services", "Utilities"],
                },
                {
                    "name": "Purchased Cooling",
                    "SubCategory": ["Electricity", "Energy Services", "Utilities"],
                },
            ],
            "Scope 3": [
                {
                    "name": "Purchased Goods & Services",
                    "SubCategory": [
                        "Agriculture/Hunting/Forestry/Fishing",
                        "Arable Farming",
                        "Building Materials",
                        "Ceramic Goods",
                        "Chemical Products",
                        "Clothing and Footwear",
                        "Cloud Computing - CPU",
                        "Cloud Computing - Memory",
                        "Cloud Computing - Networking",
                        "Cloud Computing - Storage",
                        "Consumer Goods Rental",
                        "DIY and Gardening Equipment",
                        "Domestic Services",
                        "Education",
                        "Electrical Equipment",
                        "Electronics",
                        "Equipment Rental",
                        "Fabricated Metal Products",
                        "Financial Services",
                        "Fishing/Aquaculture/Hunting",
                        "Food and Beverage Services",
                        "Food/Beverages/Tobacco",
                        "Furnishings and Household",
                        "General Retail",
                        "Glass and Glass Products",
                        "Government Activities",
                        "Health and Social Care",
                        "Health Care",
                        "Housing",
                        "Information and Communication Services",
                        "Insurance Services",
                        "Livestock Farming",
                        "Machinery",
                        "Maintenance and Repair",
                        "Manufacturing",
                        "Metals",
                        "Mined Materials",
                        "Mining",
                        "Non-profit Activities",
                        "Office Equipment",
                        "Operational Activities",
                        "Organic Products",
                        "Other Materials",
                        "Paper and Cardboard",
                        "Paper Products",
                        "Pavement and Surfacing",
                        "Personal Care and Accessories",
                        "Plastics and Rubber Products",
                        "Professional Services and Activities",
                        "Real Estate",
                        "Recreation and Culture",
                        "Social Care",
                        "Textiles",
                        "Timber and Forestry Products",
                        "Utilities",
                        "Vehicle Maintenance and Services",
                        "Vehicle Parts",
                        "Waste Management",
                        "Water Supply",
                        "Water Treatment",
                        "Wholesale Trade",
                    ],
                },
                {
                    "name": "Capital Goods",
                    "SubCategory": [
                        "Construction",
                        "Electrical Equipment",
                        "Electronics",
                        "Furnishings and Household",
                        "Health Care",
                        "Machinery",
                        "Office Equipment",
                    ],
                },
                {
                    "name": "Fuel & Energy Related Activities",
                    "SubCategory": ["Electricity", "Fuel"],
                },
                {
                    "name": "Upstream Transportation & Distribution",
                    "SubCategory": [
                        "Air Freight",
                        "Energy Services",
                        "Rail Freight",
                        "Road Freight",
                        "Sea Freight",
                        "Transport Services and Warehousing",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Waste Generated in Operations",
                    "SubCategory": [
                        "Construction Waste",
                        "Electrical Waste",
                        "Food and Organic Waste",
                        "General Waste",
                        "Glass Waste",
                        "Metal Waste",
                        "Paper and Cardboard Waste",
                        "Plastic Waste",
                        "Waste Management",
                        "Water Treatment",
                    ],
                },
                {
                    "name": "Business Travel",
                    "SubCategory": [
                        "Restaurants and Accommodation",
                        "Accommodation",
                        "Road Travel",
                        "Air Travel",
                        "Rail Travel",
                        "Sea Travel",
                        "Tickets and Passes",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Employee Commuting",
                    "SubCategory": [
                        "Homeworking",
                        "Rail Travel",
                        "Road Travel",
                        "Tickets and Passes",
                        "Vehicles",
                    ],
                },
                {
                    "name": "Upstream Leased Assets",
                    "SubCategory": ["Facility", "Housing", "Real Estate"],
                },
                {
                    "name": "Downstream Transportation & Distribution",
                    "SubCategory": [
                        "Air Freight",
                        "Energy Services",
                        "Infrastructure",
                        "Rail Freight",
                        "Road Freight",
                        "Sea Freight",
                        "Transport Services and Warehousing",
                    ],
                },
                {
                    "name": "Processing of Sold Products",
                    "SubCategory": [
                        "Cloud Computing - CPU",
                        "Cloud Computing - Memory",
                        "Cloud Computing - Networking",
                        "Cloud Computing - Storage",
                        "Information and Communication Services",
                    ],
                },
                {
                    "name": "End of Life Treatment of Sold Products",
                    "SubCategory": ["Electricity", "Fuel"],
                },
                {
                    "name": "Downstream Leased Assets",
                    "SubCategory": [
                        "Equipment Rental",
                        "Facility",
                        "Housing",
                        "Real Estate",
                    ],
                },
            ],
        }
        self.units = [
            "m2",
            "km2",
            "ft2",
            "ha",
            "containers",
            "MB",
            "GB",
            "TB",
            "m",
            "km",
            "ft",
            "mi",
            "nmi",
            "Wh",
            "kWh",
            "MWh",
            "MJ",
            "GJ",
            "TJ",
            "BTU",
            "therm",
            "MMBTU",
            "usd",
            "afn",
            "dzd",
            "ars",
            "aud",
            "bhd",
            "brl",
            "cad",
            "kyd",
            "cny",
            "dkk",
            "egp",
            "eur",
            "hkd",
            "huf",
            "isk",
            "inr",
            "iqd",
            "ils",
            "jpy",
            "lbp",
            "mxn",
            "mad",
            "nzd",
            "nok",
            "qar",
            "rub",
            "sar",
            "sgd",
            "zar",
            "krw",
            "sek",
            "chf",
            "thb",
            "twd",
            "tnd",
            "try",
            "aed",
            "gbp",
            "number of Nights",
            "numbers",
            "passengers",
            "ms",
            "s",
            "m",
            "h",
            "day",
            "year",
            "ml",
            "l",
            "m3",
            "standard_cubic_foot",
            "gallon_us",
            "bbl",
            "g",
            "kg",
            "t",
            "ton",
            "lb",
        ]

        self.units2 = ["ms", "s", "m", "h", "day", "year", "m", "km", "ft", "mi", "nmi"]
        self.months = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]

        # Build flattened mappings
        self.scope_to_categories = {}
        self.category_to_subcategories = {}

        for scope, categories in self.scope_data.items():
            cat_names = []
            for cat in categories:
                cat_name = cat["name"]
                cat_names.append(cat_name)
                self.category_to_subcategories[cat_name] = cat["SubCategory"]
            self.scope_to_categories[scope] = cat_names

    def header_body_styling(self, ws, header_row=1, body_start_row=2, body_end_row=100):
        """
        Style the Excel sheet with:
        - Frozen header row
        - Auto filter
        - Blue header with white text
        - Light green body
        - Borders to simulate gridlines
        """
        # Freeze header
        ws.freeze_panes = f"A{body_start_row}"

        # Auto filter
        max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{max_col_letter}{header_row}"

        # Styles
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        body_fill = PatternFill(
            start_color="DFF0D8", end_color="DFF0D8", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # Apply header styles
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # Apply body styles
        for row in range(body_start_row, body_end_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = body_fill
                cell.border = thin_border

    def auto_adjust_column_widths(self, ws, sample_rows=10):
        """
        Auto-adjust column widths based on max length of content
        in headers and first few data rows.
        """
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column letter (e.g. 'A', 'B', etc.)
            for cell in col[:sample_rows]:  # Check only the first few rows
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add padding
            ws.column_dimensions[column].width = adjusted_width

    def sanitize(self, name):
        # Replace non-alphanumeric characters with underscore
        name = re.sub(r"[^A-Za-z0-9_]", "_", name)
        # Ensure it starts with a letter
        if not name[0].isalpha():
            name = "N_" + name
        return name

    def add_dropdown(self, worksheet, col_letter, options, start_row=2, end_row=100):
        dv = DataValidation(
            type="list", formula1='"{}"'.format(",".join(options)), allow_blank=True
        )
        dv.error = "Select from dropdown"
        dv.errorTitle = "Invalid Entry"
        worksheet.add_data_validation(dv)
        dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        wb.properties.creator = "Sustainext"
        ws = wb.active
        ws.title = "Template"

        # Headers
        headers = [
            "Location",
            "Month",
            "Year",
            "Scope",
            "Category",
            "Subcategory",
            "Activity",
            "Quantity 1",
            "Unit 1",
            "Quantity 2",
            "Unit 2",
        ]
        ws.append(headers)
        self.header_body_styling(ws)
        self.auto_adjust_column_widths(ws)
        dv_numeric = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        dv_numeric.showErrorMessage = True
        dv_numeric.error = "Only numeric values allowed"
        dv_numeric.errorTitle = "Invalid Entry"
        ws.add_data_validation(dv_numeric)
        dv_numeric.add("H2:H100")
        dv_numeric.add("J2:J100")

        # Hidden sheet for dropdown source data
        dropdown_ws = wb.create_sheet("DropdownData")
        dropdown_ws.sheet_state = "hidden"

        # Add dropdowns for Location
        locations = list(request.user.locs.all().values_list("name", flat=True))

        # Add locations to hidden sheet (column A)
        location_col_idx = 52  # AZ
        loc_col_letter = get_column_letter(location_col_idx)
        for i, loc in enumerate(locations, start=2):
            dropdown_ws.cell(
                row=i, column=location_col_idx
            ).value = loc  # Column AZ (52) to avoid overlaps

        loc_range_end = 1 + len(locations)

        # Define named range for locations
        wb.defined_names.add(
            DefinedName(
                name="LocationList",
                attr_text=f"{quote_sheetname('DropdownData')}!${loc_col_letter}$2:${loc_col_letter}${loc_range_end}",
            )
        )
        # Location List on Column A
        for row in range(2, 101):
            formula_loc = "=LocationList"
            dv_loc = DataValidation(type="list", formula1=formula_loc, allow_blank=True)
            ws.add_data_validation(dv_loc)
            dv_loc.add(f"A{row}")  # Location column

        # Scope list in first column
        for i, scope in enumerate(self.scope_to_categories.keys(), start=2):
            dropdown_ws.cell(row=i, column=1).value = scope

        # Named ranges: scope -> category
        for col, (scope, categories) in enumerate(
            self.scope_to_categories.items(), start=2
        ):
            sanitized_scope = self.sanitize(scope)
            dropdown_ws.cell(row=1, column=col).value = sanitized_scope
            for row, cat in enumerate(categories, start=2):
                dropdown_ws.cell(row=row, column=col).value = cat
            col_letter = get_column_letter(col)
            end_row = 1 + len(categories)
            wb.defined_names.add(
                DefinedName(
                    name=sanitized_scope,
                    attr_text=f"{quote_sheetname('DropdownData')}!${col_letter}$2:${col_letter}${end_row}",
                )
            )

        # Named ranges: category -> subcategory
        for col, (cat, subcats) in enumerate(
            self.category_to_subcategories.items(), start=10
        ):
            sanitized_cat = self.sanitize(cat)
            dropdown_ws.cell(row=1, column=col).value = sanitized_cat
            for row, sub in enumerate(subcats, start=2):
                dropdown_ws.cell(row=row, column=col).value = sub
            col_letter = get_column_letter(col)
            end_row = 1 + len(subcats)
            wb.defined_names.add(
                DefinedName(
                    name=sanitized_cat,
                    attr_text=f"{quote_sheetname('DropdownData')}!${col_letter}$2:${col_letter}${end_row}",
                )
            )
        # Units (column 47 = AU), Units2 (column 48 = AV)
        units_col_idx = 47  # AU
        units2_col_idx = 48  # AV
        units_col_letter = get_column_letter(units_col_idx)
        units2_col_letter = get_column_letter(units2_col_idx)

        # Write unit values to hidden sheet
        for i, unit in enumerate(self.units, start=2):
            dropdown_ws.cell(row=i, column=units_col_idx).value = unit

        for i, unit in enumerate(self.units2, start=2):
            dropdown_ws.cell(row=i, column=units2_col_idx).value = unit

        # Define named ranges for Excel
        wb.defined_names.add(
            DefinedName(
                name="UnitList",
                attr_text=f"{quote_sheetname('DropdownData')}!${units_col_letter}$2:${units_col_letter}${1 + len(self.units)}",
            )
        )
        wb.defined_names.add(
            DefinedName(
                name="Unit2List",
                attr_text=f"{quote_sheetname('DropdownData')}!${units2_col_letter}$2:${units2_col_letter}${1 + len(self.units2)}",
            )
        )

        for row in range(2, 101):
            dv_unit = DataValidation(
                type="list", formula1="=UnitList", allow_blank=True
            )
            ws.add_data_validation(dv_unit)
            dv_unit.add(f"I{row}")

            dv_unit2 = DataValidation(
                type="list", formula1="=Unit2List", allow_blank=True
            )
            ws.add_data_validation(dv_unit2)
            dv_unit2.add(f"K{row}")

        # Static dropdowns for Month and Year

        years = [str(y) for y in range(2018, 2031)]

        self.add_dropdown(ws, "B", self.months)
        self.add_dropdown(ws, "C", years)
        self.add_dropdown(ws, "D", list(self.scope_to_categories.keys()))

        for row in range(2, 101):
            # Category depends on Scope
            formula_cat = (
                f'=INDIRECT(SUBSTITUTE(SUBSTITUTE(D{row}, " ", "_"), "&", "_"))'
            )
            dv_cat = DataValidation(type="list", formula1=formula_cat, allow_blank=True)
            ws.add_data_validation(dv_cat)
            dv_cat.add(f"E{row}")

            # Subcategory depends on Category
            formula_subcat = (
                f'=INDIRECT(SUBSTITUTE(SUBSTITUTE(E{row}, " ", "_"), "&", "_"))'
            )
            dv_subcat = DataValidation(
                type="list", formula1=formula_subcat, allow_blank=True
            )
            ws.add_data_validation(dv_subcat)
            dv_subcat.add(f"F{row}")

            # Field validation for quantity and quantity2
            ws.cell(row=row, column=8).number_format = "0.0000"
            ws.cell(row=row, column=10).number_format = "0.0000"

        # Save and respond
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=template.xlsx"
        return response
